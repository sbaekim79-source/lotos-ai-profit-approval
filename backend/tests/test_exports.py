from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app


def _login(client: TestClient, username: str, password: str) -> str:
    response = client.post(
        "/api/auth/login",
        json={"username": username, "password": password},
    )
    assert response.status_code == 200
    return response.json()["access_token"]


def _auth(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def _assert_xlsx(response, expected_sheet: str) -> None:
    assert response.status_code == 200
    assert response.content[:2] == b"PK"
    workbook = load_workbook(BytesIO(response.content))
    assert expected_sheet in workbook.sheetnames


def test_excel_exports_and_audit_log() -> None:
    with TestClient(app) as client:
        assert client.post("/api/masters/seed-defaults").status_code == 200
        admin_token = _login(client, "admin", "admin1234")
        staff_token = _login(client, "staff", "staff1234")

        save_response = client.post(
            "/api/approvals/analyze-sample-and-save/KANGKOKU_HIROBA_SI_PLUS_PLUS",
            headers=_auth(admin_token),
        )
        assert save_response.status_code == 200

        quote_response = client.post(
            "/api/quotes/generate-and-save",
            headers=_auth(admin_token),
            json={
                "customer_name": "TEST CUSTOMER",
                "trade_type": "SHIPPER",
                "partner_name": None,
                "mode": "SEA",
                "direction": "IMPORT",
                "code": "SI++",
                "pol": "BUSAN",
                "pod": "TOKYO",
                "port": "TOKYO",
                "origin": "BUSAN",
                "destination": "TOKYO",
                "container_type": "20DC",
                "container_count": 1,
                "include_customs": True,
                "include_transport": True,
            },
        )
        assert quote_response.status_code == 200

        forbidden_response = client.get(
            "/api/exports/approvals.xlsx",
            headers=_auth(staff_token),
        )
        assert forbidden_response.status_code == 403

        _assert_xlsx(
            client.get("/api/exports/approvals.xlsx", headers=_auth(admin_token)),
            "Approvals",
        )
        _assert_xlsx(
            client.get(
                "/api/exports/dashboard.xlsx",
                headers=_auth(admin_token),
                params={"work_month": "2026-06"},
            ),
            "Summary",
        )
        _assert_xlsx(
            client.get(
                "/api/exports/tariffs/transport.xlsx",
                headers=_auth(admin_token),
            ),
            "Transport Tariff",
        )
        _assert_xlsx(
            client.get(
                "/api/exports/tariffs/customs.xlsx",
                headers=_auth(admin_token),
            ),
            "Customs Tariff",
        )
        _assert_xlsx(
            client.get("/api/exports/quotes.xlsx", headers=_auth(admin_token)),
            "Quotes",
        )
        _assert_xlsx(
            client.get(
                "/api/exports/productivity.xlsx",
                headers=_auth(admin_token),
                params={"start_month": "2026-01", "end_month": "2026-12"},
            ),
            "Productivity",
        )

        audit_response = client.get(
            "/api/audit-logs",
            params={"action": "EXPORT_EXCEL"},
        )
        assert audit_response.status_code == 200
        audit_rows = audit_response.json()
        assert len(audit_rows) >= 6
        assert {row["entity_type"] for row in audit_rows} >= {
            "APPROVALS",
            "DASHBOARD",
            "TARIFF_TRANSPORT",
            "TARIFF_CUSTOMS",
            "QUOTES",
            "PRODUCTIVITY",
        }
