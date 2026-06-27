from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import settings


EXPORT_DIR = settings.export_dir
EXPORT_DIR.mkdir(parents=True, exist_ok=True)


def _cell_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return value


def _safe_sheet_name(name: str) -> str:
    invalid_chars = ["\\", "/", "*", "?", ":", "[", "]"]
    safe_name = name
    for char in invalid_chars:
        safe_name = safe_name.replace(char, " ")
    return safe_name[:31] or "Sheet"


def create_excel_file(file_name: str, sheets: dict[str, list[dict[str, Any]]]) -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name, rows in sheets.items():
        worksheet = workbook.create_sheet(_safe_sheet_name(sheet_name))
        headers = list(rows[0].keys()) if rows else ["No Data"]
        worksheet.append(headers)

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")

        if rows:
            for row in rows:
                worksheet.append([_cell_value(row.get(header)) for header in headers])

        for column_index, header in enumerate(headers, start=1):
            max_length = len(str(header))
            for cell in worksheet.iter_cols(
                min_col=column_index,
                max_col=column_index,
                min_row=2,
                values_only=True,
            ):
                for value in cell:
                    if value is not None:
                        max_length = max(max_length, len(str(value)))
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(
                max_length + 2,
                48,
            )

    path = EXPORT_DIR / file_name
    workbook.save(path)
    return path
